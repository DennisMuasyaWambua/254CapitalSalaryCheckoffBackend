"""
Export generators for Excel and PDF files.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML, CSS
from django.template.loader import render_to_string
from django.conf import settings
from io import BytesIO
from datetime import date, datetime
from decimal import Decimal
import logging

logger = logging.getLogger(__name__)


def generate_deduction_list_excel(
    month: int,
    year: int,
    employer_name: str,
    deductions: list
) -> BytesIO:
    """
    Generate Excel file with deduction list.

    Args:
        month: Month number (1-12)
        year: Year
        employer_name: Employer name
        deductions: List of deduction records

    Returns:
        BytesIO object with Excel file
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Deduction List'

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='3F2A56', end_color='3F2A56', fill_type='solid')
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f'{employer_name} - Loan Deduction List'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Period
        ws.merge_cells('A2:F2')
        period_name = date(year, month, 1).strftime('%B %Y')
        ws['A2'] = f'Period: {period_name}'
        ws['A2'].alignment = Alignment(horizontal='center')

        # Empty row
        ws.append([])

        # Headers
        headers = ['Employee Name', 'Employee ID', 'Loan Number', 'Deduction Amount', 'Deduction Tag', 'Notes']
        ws.append(headers)

        header_row = ws[4]
        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Data rows
        for deduction in deductions:
            row = [
                deduction['employee_name'],
                deduction['employee_id'],
                deduction['loan_number'],
                float(deduction['amount']),
                deduction['tag'],  # "This Month" or "Next Month"
                deduction.get('notes', '')
            ]
            ws.append(row)

            # Apply currency format to amount column
            amount_cell = ws.cell(row=ws.max_row, column=4)
            amount_cell.number_format = '#,##0.00'
            amount_cell.border = border

            # Apply color to tag column
            tag_cell = ws.cell(row=ws.max_row, column=5)
            if deduction['tag'] == 'This Month':
                tag_cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # Light green
            else:
                tag_cell.fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')  # Light orange

            # Apply borders to all cells in row
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(vertical='center')

        # Summary row
        ws.append([])
        summary_row = ws.max_row + 1
        ws[f'A{summary_row}'] = 'TOTAL'
        ws[f'A{summary_row}'].font = Font(bold=True)

        total_amount = sum(deduction['amount'] for deduction in deductions)
        ws[f'D{summary_row}'] = float(total_amount)
        ws[f'D{summary_row}'].number_format = '#,##0.00'
        ws[f'D{summary_row}'].font = Font(bold=True)

        # Adjust column widths
        column_widths = [30, 15, 20, 18, 18, 30]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f'Generated deduction Excel for {employer_name} - {period_name}')

        return output

    except Exception as e:
        logger.error(f'Failed to generate deduction Excel: {str(e)}')
        raise


def generate_repayment_schedule_pdf(loan_application) -> BytesIO:
    """
    Generate PDF repayment schedule.

    Args:
        loan_application: LoanApplication instance

    Returns:
        BytesIO object with PDF file
    """
    try:
        # Get repayment schedule
        schedule = loan_application.repayment_schedule.all().order_by('installment_number')

        # Prepare context
        context = {
            'loan': loan_application,
            'employer': loan_application.employer,
            'employee': loan_application.employee,
            'schedule': schedule,
            'generated_date': datetime.now().strftime('%d %B %Y'),
        }

        # HTML template
        html_string = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        @page {
            size: A4;
            margin: 1.5cm;
        }
        body {
            font-family: 'Arial', sans-serif;
            font-size: 10pt;
            color: #333;
        }
        .header {
            text-align: center;
            margin-bottom: 30px;
            border-bottom: 3px solid #3F2A56;
            padding-bottom: 15px;
        }
        .header h1 {
            color: #3F2A56;
            margin: 0;
            font-size: 20pt;
        }
        .header h2 {
            color: #00BCD4;
            margin: 5px 0;
            font-size: 16pt;
        }
        .info-section {
            margin-bottom: 25px;
            display: table;
            width: 100%;
        }
        .info-row {
            display: table-row;
        }
        .info-label {
            display: table-cell;
            font-weight: bold;
            width: 150px;
            padding: 5px 0;
        }
        .info-value {
            display: table-cell;
            padding: 5px 0;
        }
        .summary-box {
            background-color: #F5F5F5;
            border: 2px solid #3F2A56;
            border-radius: 5px;
            padding: 15px;
            margin: 20px 0;
        }
        .summary-box h3 {
            color: #3F2A56;
            margin-top: 0;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }
        th {
            background-color: #3F2A56;
            color: white;
            padding: 10px;
            text-align: left;
            font-weight: bold;
        }
        td {
            padding: 8px 10px;
            border-bottom: 1px solid #ddd;
        }
        tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        tr.first-deduction {
            background-color: #C8E6C9 !important;
            font-weight: bold;
        }
        .amount {
            text-align: right;
        }
        .footer {
            margin-top: 50px;
            page-break-inside: avoid;
        }
        .signature-section {
            display: table;
            width: 100%;
            margin-top: 40px;
        }
        .signature-box {
            display: table-cell;
            width: 45%;
            border-top: 1px solid #333;
            padding-top: 5px;
        }
        .signature-spacer {
            display: table-cell;
            width: 10%;
        }
        .notes {
            font-size: 9pt;
            color: #666;
            margin-top: 20px;
        }
    </style>
</head>
<body>
    <div class="header">
        <h1>254 CAPITAL</h1>
        <h2>Loan Repayment Schedule</h2>
    </div>

    <div class="info-section">
        <div class="info-row">
            <div class="info-label">Loan Number:</div>
            <div class="info-value">{{ loan.application_number }}</div>
        </div>
        <div class="info-row">
            <div class="info-label">Employee Name:</div>
            <div class="info-value">{{ employee.get_full_name }}</div>
        </div>
        <div class="info-row">
            <div class="info-label">Employee ID:</div>
            <div class="info-value">{{ employee.employee_profile.employee_id }}</div>
        </div>
        <div class="info-row">
            <div class="info-label">Employer:</div>
            <div class="info-value">{{ employer.name }}</div>
        </div>
        <div class="info-row">
            <div class="info-label">Disbursement Date:</div>
            <div class="info-value">{{ loan.disbursement_date|date:"d F Y" }}</div>
        </div>
    </div>

    <div class="summary-box">
        <h3>Loan Summary</h3>
        <div class="info-section">
            <div class="info-row">
                <div class="info-label">Principal Amount:</div>
                <div class="info-value">KES {{ loan.principal_amount|floatformat:2|intcomma }}</div>
            </div>
            <div class="info-row">
                <div class="info-label">Interest Rate:</div>
                <div class="info-value">{{ loan.interest_rate|floatformat:2 }}% Flat</div>
            </div>
            <div class="info-row">
                <div class="info-label">Repayment Period:</div>
                <div class="info-value">{{ loan.repayment_months }} Months</div>
            </div>
            <div class="info-row">
                <div class="info-label">Total Repayment:</div>
                <div class="info-value"><strong>KES {{ loan.total_repayment|floatformat:2|intcomma }}</strong></div>
            </div>
            <div class="info-row">
                <div class="info-label">Monthly Deduction:</div>
                <div class="info-value"><strong>KES {{ loan.monthly_deduction|floatformat:2|intcomma }}</strong></div>
            </div>
            <div class="info-row">
                <div class="info-label">First Deduction:</div>
                <div class="info-value">{{ loan.first_deduction_date|date:"d F Y" }}</div>
            </div>
        </div>
    </div>

    <h3>Repayment Schedule</h3>
    <table>
        <thead>
            <tr>
                <th>Installment</th>
                <th>Due Date</th>
                <th class="amount">Deduction Amount</th>
                <th class="amount">Running Balance</th>
                <th>Status</th>
            </tr>
        </thead>
        <tbody>
            {% for item in schedule %}
            <tr{% if item.is_first_deduction %} class="first-deduction"{% endif %}>
                <td>{{ item.installment_number }}</td>
                <td>{{ item.due_date|date:"d F Y" }}</td>
                <td class="amount">KES {{ item.amount|floatformat:2|intcomma }}</td>
                <td class="amount">KES {{ item.running_balance|floatformat:2|intcomma }}</td>
                <td>{% if item.is_paid %}Paid{% else %}Pending{% endif %}</td>
            </tr>
            {% endfor %}
        </tbody>
    </table>

    <div class="notes">
        <p><strong>Notes:</strong></p>
        <ul>
            <li>Deductions will be made on the 25th of each month from your salary.</li>
            <li>The first deduction is highlighted in green.</li>
            <li>This is an interest-bearing loan at {{ loan.interest_rate|floatformat:2 }}% flat rate.</li>
            <li>Please ensure sufficient salary balance for deductions.</li>
        </ul>
    </div>

    <div class="footer">
        <div class="signature-section">
            <div class="signature-box">
                <strong>Employee Signature</strong><br>
                Date: _________________
            </div>
            <div class="signature-spacer"></div>
            <div class="signature-box">
                <strong>254 Capital Representative</strong><br>
                Date: _________________
            </div>
        </div>

        <p style="text-align: center; margin-top: 30px; font-size: 9pt; color: #666;">
            Generated on {{ generated_date }}<br>
            254 Capital - Empowering Financial Freedom
        </p>
    </div>
</body>
</html>
        """

        # Render HTML with context
        from django.template import Context, Template
        template = Template(html_string)
        html_rendered = template.render(Context(context))

        # Generate PDF
        pdf_file = HTML(string=html_rendered).write_pdf()

        output = BytesIO(pdf_file)
        output.seek(0)

        logger.info(f'Generated repayment schedule PDF for {loan_application.application_number}')

        return output

    except Exception as e:
        logger.error(f'Failed to generate repayment PDF: {str(e)}')
        raise


def generate_loan_book_report_data(employer_id=None):
    """
    Generate loan book report data for charts.

    Args:
        employer_id: Optional employer filter

    Returns:
        Dict with report data
    """
    try:
        from apps.loans.models import LoanApplication
        from django.db.models import Sum, Count, Q

        queryset = LoanApplication.objects.all()

        if employer_id:
            queryset = queryset.filter(employer_id=employer_id)

        # Total disbursed
        total_disbursed = queryset.filter(
            status=LoanApplication.Status.DISBURSED
        ).aggregate(
            total=Sum('principal_amount'),
            count=Count('id')
        )

        # Total outstanding
        total_outstanding = queryset.filter(
            status=LoanApplication.Status.DISBURSED
        ).aggregate(
            total=Sum('total_repayment')
        )

        # Status breakdown
        status_breakdown = []
        for status_choice in LoanApplication.Status.choices:
            status_value = status_choice[0]
            status_label = status_choice[1]

            count = queryset.filter(status=status_value).count()
            total = queryset.filter(status=status_value).aggregate(
                total=Sum('principal_amount')
            )['total'] or Decimal('0.00')

            status_breakdown.append({
                'status': status_value,
                'label': status_label,
                'count': count,
                'total': float(total)
            })

        # By employer (if not filtering by employer)
        by_employer = []
        if not employer_id:
            from apps.employers.models import Employer

            employers = Employer.objects.all()
            for employer in employers:
                employer_loans = queryset.filter(employer=employer)

                active_count = employer_loans.filter(status=LoanApplication.Status.DISBURSED).count()
                active_total = employer_loans.filter(
                    status=LoanApplication.Status.DISBURSED
                ).aggregate(total=Sum('principal_amount'))['total'] or Decimal('0.00')

                by_employer.append({
                    'employer_id': str(employer.id),
                    'employer_name': employer.name,
                    'active_loans': active_count,
                    'total_disbursed': float(active_total)
                })

        return {
            'total_disbursed_amount': float(total_disbursed['total'] or 0),
            'total_disbursed_count': total_disbursed['count'],
            'total_outstanding_amount': float(total_outstanding['total'] or 0),
            'status_breakdown': status_breakdown,
            'by_employer': by_employer
        }

    except Exception as e:
        logger.error(f'Failed to generate loan book report: {str(e)}')
        raise
