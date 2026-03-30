"""
Views for client management endpoints.
"""

from rest_framework import viewsets, status, generics
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging

from .models import ExistingClient
from .serializers import (
    ExistingClientSerializer,
    ExistingClientBulkUploadSerializer,
    ExistingClientApprovalSerializer,
    BulkApprovalSerializer,
)
from common.email_service import send_email, send_internal_alert

logger = logging.getLogger(__name__)


class ExistingClientViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing existing client records.

    Provides endpoints for:
    - Creating manual client entries
    - Bulk upload from Excel/CSV
    - Listing and filtering clients
    - Approving/rejecting clients
    """

    queryset = ExistingClient.objects.select_related('employer').all()
    serializer_class = ExistingClientSerializer
    permission_classes = [IsAuthenticated]
    filterset_fields = ['approval_status', 'loan_status', 'employer']
    search_fields = ['full_name', 'national_id', 'mobile', 'employee_id']
    ordering_fields = ['created_at', 'full_name', 'outstanding_balance']

    def get_permissions(self):
        """
        Override permissions for specific actions.
        Allow unauthenticated access to template download.
        """
        if self.action == 'upload_template':
            return [AllowAny()]
        return super().get_permissions()

    def get_queryset(self):
        """Filter queryset based on user role."""
        queryset = super().get_queryset()

        # Admins see all clients
        if self.request.user.role == 'admin':
            return queryset

        # HR users see only their employer's clients
        if self.request.user.role == 'hr':
            return queryset.filter(employer=self.request.user.employer)

        # Employees should not access this endpoint
        return queryset.none()

    @action(detail=False, methods=['post'])
    def manual(self, request):
        """
        Create a manual existing client entry.

        POST /api/v1/clients/manual/
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Set entered_by to current user's name
        client = serializer.save(entered_by=request.user.get_full_name())

        # Send internal alert
        try:
            alert_message = f"""
            <p><strong>New Existing Client Record Created (Manual Entry)</strong></p>
            <ul>
                <li><strong>Client:</strong> {client.full_name}</li>
                <li><strong>National ID:</strong> {client.national_id}</li>
                <li><strong>Mobile:</strong> {client.mobile}</li>
                <li><strong>Employer:</strong> {client.employer.name}</li>
                <li><strong>Loan Amount:</strong> KES {client.loan_amount:,.2f}</li>
                <li><strong>Entered by:</strong> {request.user.get_full_name()}</li>
            </ul>
            """
            send_internal_alert(
                subject=f'New Client Record - {client.full_name}',
                message=alert_message,
                alert_type='info'
            )
        except Exception as e:
            logger.error(f'Failed to send internal alert: {str(e)}')

        return Response({
            'detail': 'Client record created successfully',
            'client': ExistingClientSerializer(client).data
        }, status=status.HTTP_201_CREATED)

    def _get_employer_id(self, employer_value):
        """
        Helper method to get employer ID from either name or UUID.

        Args:
            employer_value: Either employer name (str) or employer UUID (str)

        Returns:
            str: Employer UUID

        Raises:
            ValueError: If employer not found
        """
        from apps.employers.models import Employer
        from uuid import UUID

        if not employer_value or pd.isna(employer_value):
            raise ValueError("Employer is required")

        employer_value = str(employer_value).strip()

        # Try to parse as UUID first
        try:
            UUID(employer_value)
            # It's a valid UUID, verify it exists
            employer = Employer.objects.filter(id=employer_value, is_active=True).first()
            if employer:
                return employer_value
            else:
                raise ValueError(f"Employer with ID '{employer_value}' not found or inactive")
        except (ValueError, AttributeError):
            # Not a UUID, treat as employer name
            employer = Employer.objects.filter(name__iexact=employer_value, is_active=True).first()
            if employer:
                return str(employer.id)
            else:
                raise ValueError(f"Employer '{employer_value}' not found. Please check the 'Employer Reference' sheet.")

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        """
        Bulk upload client records from Excel/CSV file.
        Accepts employer names or IDs.

        POST /api/v1/clients/bulk-upload/
        """
        upload_serializer = ExistingClientBulkUploadSerializer(data=request.data)
        upload_serializer.is_valid(raise_exception=True)

        file = request.FILES['file']

        try:
            # Read file into pandas DataFrame
            if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
                df = pd.read_excel(file)
            elif file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                return Response({
                    'error': 'Invalid file format'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Process and validate rows
            valid_clients = []
            errors = []

            for index, row in df.iterrows():
                try:
                    # Get employer value (can be name or ID)
                    employer_value = (
                        row.get('Employer Name *') or
                        row.get('Employer Name') or
                        row.get('Employer ID') or
                        row.get('employer_name') or
                        row.get('employer_id')
                    )

                    # Resolve employer ID
                    employer_id = self._get_employer_id(employer_value)

                    # Map DataFrame columns to model fields (matching new template headers)
                    client_data = {
                        'full_name': (
                            row.get('Full Name *') or
                            row.get('Full Name') or
                            row.get('full_name')
                        ),
                        'national_id': str(
                            row.get('National ID Number *') or
                            row.get('National ID Number') or
                            row.get('National ID') or
                            row.get('national_id')
                        ),
                        'mobile': str(
                            row.get('Mobile Number *') or
                            row.get('Mobile Number') or
                            row.get('Mobile') or
                            row.get('mobile')
                        ),
                        'email': (
                            row.get('Email Address') or
                            row.get('Email') or
                            row.get('email', '')
                        ),
                        'employer': employer_id,
                        'employee_id': (
                            row.get('Employee ID') or
                            row.get('employee_id', '')
                        ),
                        'loan_amount': float(
                            row.get('Loan Amount (KES) *') or
                            row.get('Loan Amount (KES)') or
                            row.get('Loan Amount') or
                            row.get('loan_amount')
                        ),
                        'interest_rate': float(
                            row.get('Interest Rate (%) *') or
                            row.get('Interest Rate (%)') or
                            row.get('Interest Rate') or
                            row.get('interest_rate')
                        ),
                        'start_date': pd.to_datetime(
                            row.get('Loan Start Date *') or
                            row.get('Loan Start Date') or
                            row.get('Start Date') or
                            row.get('start_date')
                        ).date(),
                        'repayment_period': int(
                            row.get('Repayment Period (Months) *') or
                            row.get('Repayment Period (Months)') or
                            row.get('Repayment Period') or
                            row.get('repayment_period')
                        ),
                        'disbursement_date': pd.to_datetime(
                            row.get('Disbursement Date *') or
                            row.get('Disbursement Date') or
                            row.get('disbursement_date')
                        ).date(),
                        'disbursement_method': str(
                            row.get('Disbursement Method *') or
                            row.get('Disbursement Method') or
                            row.get('disbursement_method', 'mpesa')
                        ).lower().strip(),
                        'amount_paid': float(
                            row.get('Amount Paid to Date (KES)') or
                            row.get('Amount Paid') or
                            row.get('amount_paid', 0)
                        ),
                        'loan_status': (
                            row.get('Loan Status') or
                            row.get('loan_status', 'Active')
                        ),
                    }

                    # Skip rows that appear to be instructions or empty
                    if not client_data['full_name'] or pd.isna(client_data['full_name']):
                        continue
                    if str(client_data['full_name']).startswith('INSTRUCTIONS') or str(client_data['full_name']).startswith('1.'):
                        continue

                    # Validate and create client
                    serializer = ExistingClientSerializer(data=client_data)
                    serializer.is_valid(raise_exception=True)
                    client = serializer.save(entered_by=request.user.get_full_name())
                    valid_clients.append(str(client.id))

                except Exception as e:
                    errors.append({
                        'row': index + 2,  # +2 because index starts at 0 and we have header row
                        'error': str(e)
                    })

            # Send internal alert for bulk upload
            if valid_clients:
                try:
                    alert_message = f"""
                    <p><strong>Bulk Client Upload Completed</strong></p>
                    <ul>
                        <li><strong>Total Rows:</strong> {len(df)}</li>
                        <li><strong>Successful:</strong> {len(valid_clients)}</li>
                        <li><strong>Failed:</strong> {len(errors)}</li>
                        <li><strong>Uploaded by:</strong> {request.user.get_full_name()}</li>
                        <li><strong>File:</strong> {file.name}</li>
                    </ul>
                    """
                    send_internal_alert(
                        subject=f'Bulk Client Upload - {len(valid_clients)} records added',
                        message=alert_message,
                        alert_type='info'
                    )
                except Exception as e:
                    logger.error(f'Failed to send internal alert: {str(e)}')

            return Response({
                'message': 'Bulk upload processed',
                'total_rows': len(df),
                'successful': len(valid_clients),
                'failed': len(errors),
                'valid_client_ids': valid_clients,
                'errors': errors[:50]  # Limit errors to first 50
            }, status=status.HTTP_201_CREATED if valid_clients else status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response({
                'error': f'Failed to process file: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def validate_upload(self, request):
        """
        Validate bulk upload file without importing.
        Provides preview with validation status for each row.

        POST /api/v1/clients/validate/
        """
        upload_serializer = ExistingClientBulkUploadSerializer(data=request.data)
        upload_serializer.is_valid(raise_exception=True)

        file = request.FILES['file']

        try:
            # Read file
            if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
                df = pd.read_excel(file)
            elif file.name.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                return Response({
                    'error': 'Invalid file format'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Validate rows
            preview_data = []

            for index, row in df.iterrows():
                try:
                    # Get employer value (can be name or ID)
                    employer_value = (
                        row.get('Employer Name *') or
                        row.get('Employer Name') or
                        row.get('Employer ID') or
                        row.get('employer_name') or
                        row.get('employer_id')
                    )

                    # Get full name
                    full_name = (
                        row.get('Full Name *') or
                        row.get('Full Name') or
                        row.get('full_name')
                    )

                    # Skip instruction rows or empty rows
                    if not full_name or pd.isna(full_name):
                        continue
                    if str(full_name).startswith('INSTRUCTIONS') or str(full_name).startswith('1.'):
                        continue

                    # Resolve employer ID
                    employer_id = self._get_employer_id(employer_value)

                    client_data = {
                        'full_name': full_name,
                        'national_id': str(
                            row.get('National ID Number *') or
                            row.get('National ID Number') or
                            row.get('National ID') or
                            row.get('national_id')
                        ),
                        'mobile': str(
                            row.get('Mobile Number *') or
                            row.get('Mobile Number') or
                            row.get('Mobile') or
                            row.get('mobile')
                        ),
                        'email': (
                            row.get('Email Address') or
                            row.get('Email') or
                            row.get('email', '')
                        ),
                        'employer': employer_id,
                        'employee_id': (
                            row.get('Employee ID') or
                            row.get('employee_id', '')
                        ),
                        'loan_amount': float(
                            row.get('Loan Amount (KES) *') or
                            row.get('Loan Amount (KES)') or
                            row.get('Loan Amount') or
                            row.get('loan_amount')
                        ),
                        'interest_rate': float(
                            row.get('Interest Rate (%) *') or
                            row.get('Interest Rate (%)') or
                            row.get('Interest Rate') or
                            row.get('interest_rate')
                        ),
                        'start_date': pd.to_datetime(
                            row.get('Loan Start Date *') or
                            row.get('Loan Start Date') or
                            row.get('Start Date') or
                            row.get('start_date')
                        ).date(),
                        'repayment_period': int(
                            row.get('Repayment Period (Months) *') or
                            row.get('Repayment Period (Months)') or
                            row.get('Repayment Period') or
                            row.get('repayment_period')
                        ),
                        'disbursement_date': pd.to_datetime(
                            row.get('Disbursement Date *') or
                            row.get('Disbursement Date') or
                            row.get('disbursement_date')
                        ).date(),
                        'disbursement_method': str(
                            row.get('Disbursement Method *') or
                            row.get('Disbursement Method') or
                            row.get('disbursement_method', 'mpesa')
                        ).lower().strip(),
                        'amount_paid': float(
                            row.get('Amount Paid to Date (KES)') or
                            row.get('Amount Paid') or
                            row.get('amount_paid', 0)
                        ),
                        'loan_status': (
                            row.get('Loan Status') or
                            row.get('loan_status', 'Active')
                        ),
                    }

                    # Validate with serializer
                    serializer = ExistingClientSerializer(data=client_data)
                    if serializer.is_valid():
                        preview_data.append({
                            'row_number': index + 2,
                            'name': client_data['full_name'],
                            'national_id': client_data['national_id'],
                            'mobile': client_data['mobile'],
                            'employer': str(employer_value),
                            'loan_amount': client_data['loan_amount'],
                            'status': 'valid',
                            'issue': None
                        })
                    else:
                        error_messages = []
                        for field, errors in serializer.errors.items():
                            error_messages.append(f"{field}: {', '.join(errors)}")
                        preview_data.append({
                            'row_number': index + 2,
                            'name': str(full_name),
                            'national_id': str(client_data.get('national_id', '')),
                            'mobile': str(client_data.get('mobile', '')),
                            'employer': str(employer_value),
                            'loan_amount': client_data.get('loan_amount', 0),
                            'status': 'error',
                            'issue': '; '.join(error_messages)
                        })

                except Exception as e:
                    preview_data.append({
                        'row_number': index + 2,
                        'name': str(full_name) if 'full_name' in locals() and full_name else 'Unknown',
                        'national_id': '',
                        'mobile': '',
                        'employer': str(employer_value) if 'employer_value' in locals() else '',
                        'loan_amount': 0,
                        'status': 'error',
                        'issue': str(e)
                    })

            valid_count = sum(1 for r in preview_data if r['status'] == 'valid')
            invalid_count = len(preview_data) - valid_count

            return Response({
                'total_rows': len(df),
                'valid_rows': valid_count,
                'invalid_rows': invalid_count,
                'preview': preview_data[:100]  # Limit to first 100
            })

        except Exception as e:
            return Response({
                'error': f'Failed to validate file: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], permission_classes=[AllowAny], authentication_classes=[])
    def upload_template(self, request):
        """
        Generate and return Excel template file with employer reference.

        GET /api/v1/clients/upload-template/
        """
        from apps.employers.models import Employer
        from openpyxl.styles import Font, PatternFill, Alignment

        # Create workbook
        wb = Workbook()

        # Sheet 1: Client Data Template
        ws1 = wb.active
        ws1.title = "Client Data"

        # Define headers (matching frontend form order exactly)
        headers = [
            'Full Name *',
            'National ID Number *',
            'Mobile Number *',
            'Email Address',
            'Employer Name *',
            'Employee ID',
            'Loan Amount (KES) *',
            'Interest Rate (%) *',
            'Loan Start Date *',
            'Repayment Period (Months) *',
            'Disbursement Date *',
            'Disbursement Method *',
            'Amount Paid to Date (KES)',
            'Loan Status'
        ]

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write headers with styling
        for col, header in enumerate(headers, start=1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Adjust column widths
        column_widths = [20, 18, 15, 25, 25, 15, 18, 15, 18, 22, 18, 20, 22, 15]
        for col, width in enumerate(column_widths, start=1):
            ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = width

        # Add sample data rows
        sample_data_rows = [
            [
                'John Kamau',
                '12345678',
                '254712345678',
                'john@example.com',
                'Safaricom PLC',  # Use employer name instead of ID
                'EMP-1234',
                100000,
                5,
                '2026-01-01',
                6,
                '2026-01-05',
                'mpesa',
                0,
                'Active'
            ],
            [
                'Jane Wanjiku',
                '87654321',
                '254723456789',
                'jane@example.com',
                'Kenya Power',
                'EMP-5678',
                150000,
                5,
                '2026-02-01',
                12,
                '2026-02-05',
                'bank',
                25000,
                'Active'
            ]
        ]

        for row_idx, sample_data in enumerate(sample_data_rows, start=2):
            for col, value in enumerate(sample_data, start=1):
                ws1.cell(row=row_idx, column=col, value=value)

        # Add instructions row
        ws1.cell(row=4, column=1, value="INSTRUCTIONS:")
        ws1.cell(row=4, column=1).font = Font(bold=True, color="FF0000")

        instructions = [
            "1. Fields marked with * are required",
            "2. Use exact employer name from 'Employer Reference' sheet",
            "3. Dates must be in YYYY-MM-DD format (e.g., 2026-01-15)",
            "4. Disbursement Method: mpesa, bank, or cash",
            "5. Loan Status: Active, Fully Paid, Defaulted, or Restructured",
            "6. Delete sample rows and this instruction section before uploading"
        ]

        for idx, instruction in enumerate(instructions, start=5):
            ws1.cell(row=idx, column=1, value=instruction)
            ws1.cell(row=idx, column=1).font = Font(italic=True, color="666666")

        # Sheet 2: Employer Reference
        ws2 = wb.create_sheet(title="Employer Reference")

        # Employer reference headers
        ws2.cell(row=1, column=1, value="Employer Name")
        ws2.cell(row=1, column=2, value="Employer ID (for reference)")
        ws2.cell(row=1, column=1).font = header_font
        ws2.cell(row=1, column=2).font = header_font
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=2).fill = header_fill
        ws2.cell(row=1, column=1).alignment = header_alignment
        ws2.cell(row=1, column=2).alignment = header_alignment

        # Set column widths
        ws2.column_dimensions['A'].width = 35
        ws2.column_dimensions['B'].width = 40

        # Add active employers
        employers = Employer.objects.filter(is_active=True).order_by('name')
        for idx, employer in enumerate(employers, start=2):
            ws2.cell(row=idx, column=1, value=employer.name)
            ws2.cell(row=idx, column=2, value=str(employer.id))

        # Add note at the bottom
        note_row = len(employers) + 3
        ws2.cell(row=note_row, column=1, value="NOTE: Copy the exact employer name to the 'Client Data' sheet")
        ws2.cell(row=note_row, column=1).font = Font(bold=True, color="0000FF")

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from django.http import HttpResponse
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=client_upload_template.xlsx'

        return response

    @action(detail=False, methods=['get'])
    def pending(self, request):
        """
        List all pending client approvals.

        GET /api/v1/clients/pending/
        """
        pending_clients = self.get_queryset().filter(approval_status='pending')
        page = self.paginate_queryset(pending_clients)

        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(pending_clients, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """
        Approve a client record.

        POST /api/v1/clients/{id}/approve/
        """
        client = self.get_object()

        if client.approval_status != 'pending':
            return Response({
                'error': 'Client is not in pending status'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Update approval status
        client.approval_status = 'approved'
        client.save()

        # TODO: Create employee account if needed
        # TODO: Send SMS notification

        # Send email notification if client has email
        if client.email:
            try:
                subject = 'Welcome to 254 Capital - Loan Record Approved'
                body_html = f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="UTF-8">
                    <style>
                        body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                        .header {{ background-color: #27ae60; color: white; padding: 20px; text-align: center; }}
                        .content {{ padding: 20px; background-color: #f9f9f9; }}
                        .footer {{ padding: 20px; text-align: center; font-size: 12px; color: #666; }}
                        .info-box {{ background-color: #e8f8f5; border-left: 4px solid #27ae60; padding: 15px; margin: 15px 0; }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <div class="header">
                            <h1>Loan Record Approved</h1>
                        </div>
                        <div class="content">
                            <h2>Hello {client.full_name},</h2>
                            <p>Your existing loan record with 254 Capital has been approved and is now active in our system.</p>

                            <div class="info-box">
                                <p><strong>Loan Details:</strong></p>
                                <ul>
                                    <li><strong>Loan Amount:</strong> KES {client.loan_amount:,.2f}</li>
                                    <li><strong>Repayment Period:</strong> {client.repayment_period} months</li>
                                    <li><strong>Monthly Deduction:</strong> KES {client.monthly_deduction:,.2f}</li>
                                    <li><strong>Outstanding Balance:</strong> KES {client.outstanding_balance:,.2f}</li>
                                    <li><strong>Employer:</strong> {client.employer.name}</li>
                                </ul>
                            </div>

                            <p>If you have any questions, please contact us.</p>

                            <p>Best regards,<br>
                            <strong>254 Capital Team</strong></p>
                        </div>
                        <div class="footer">
                            <p>&copy; 2026 254 Capital. All rights reserved.</p>
                        </div>
                    </div>
                </body>
                </html>
                """
                send_email(client.email, subject, body_html, cc_address='david.muema@254-capital.com')
                logger.info(f'Approval email sent to client {client.id}')
            except Exception as e:
                logger.error(f'Failed to send approval email: {str(e)}')

        # Send internal alert
        try:
            alert_message = f"""
            <p><strong>Existing Client Record Approved</strong></p>
            <ul>
                <li><strong>Client:</strong> {client.full_name}</li>
                <li><strong>National ID:</strong> {client.national_id}</li>
                <li><strong>Mobile:</strong> {client.mobile}</li>
                <li><strong>Employer:</strong> {client.employer.name}</li>
                <li><strong>Loan Amount:</strong> KES {client.loan_amount:,.2f}</li>
                <li><strong>Outstanding Balance:</strong> KES {client.outstanding_balance:,.2f}</li>
                <li><strong>Approved by:</strong> {request.user.get_full_name()}</li>
            </ul>
            """
            send_internal_alert(
                subject=f'Client Approved - {client.full_name}',
                message=alert_message,
                alert_type='success'
            )
        except Exception as e:
            logger.error(f'Failed to send internal alert: {str(e)}')

        return Response({
            'detail': 'Client approved successfully',
            'client': ExistingClientSerializer(client).data
        })

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """
        Reject a client record.

        POST /api/v1/clients/{id}/reject/
        """
        client = self.get_object()

        if client.approval_status != 'pending':
            return Response({
                'error': 'Client is not in pending status'
            }, status=status.HTTP_400_BAD_REQUEST)

        serializer = ExistingClientApprovalSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Update approval status and rejection reason
        client.approval_status = 'rejected'
        client.rejection_reason = serializer.validated_data.get('rejection_reason', '')
        client.save()

        # Send email notification if client has email
        if client.email:
            try:
                subject = '254 Capital - Loan Record Review Update'
                body_html = f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="UTF-8">
                    <style>
                        body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                        .header {{ background-color: #e74c3c; color: white; padding: 20px; text-align: center; }}
                        .content {{ padding: 20px; background-color: #f9f9f9; }}
                        .footer {{ padding: 20px; text-align: center; font-size: 12px; color: #666; }}
                        .warning {{ background-color: #ffebee; border-left: 4px solid #e74c3c; padding: 15px; margin: 15px 0; }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <div class="header">
                            <h1>Loan Record Review Update</h1>
                        </div>
                        <div class="content">
                            <h2>Hello {client.full_name},</h2>
                            <p>We regret to inform you that your loan record submission requires further review.</p>

                            <div class="warning">
                                <p><strong>Reason for Review:</strong></p>
                                <p>{client.rejection_reason or 'Please contact us for more information.'}</p>
                            </div>

                            <p>Please contact our support team for more information or to resubmit your application.</p>

                            <p>Best regards,<br>
                            <strong>254 Capital Team</strong></p>
                        </div>
                        <div class="footer">
                            <p>&copy; 2026 254 Capital. All rights reserved.</p>
                        </div>
                    </div>
                </body>
                </html>
                """
                send_email(client.email, subject, body_html, cc_address='david.muema@254-capital.com')
                logger.info(f'Rejection email sent to client {client.id}')
            except Exception as e:
                logger.error(f'Failed to send rejection email: {str(e)}')

        # Send internal alert
        try:
            alert_message = f"""
            <p><strong>Existing Client Record Rejected</strong></p>
            <ul>
                <li><strong>Client:</strong> {client.full_name}</li>
                <li><strong>National ID:</strong> {client.national_id}</li>
                <li><strong>Employer:</strong> {client.employer.name}</li>
                <li><strong>Rejection Reason:</strong> {client.rejection_reason or 'Not specified'}</li>
                <li><strong>Rejected by:</strong> {request.user.get_full_name()}</li>
            </ul>
            """
            send_internal_alert(
                subject=f'Client Rejected - {client.full_name}',
                message=alert_message,
                alert_type='warning'
            )
        except Exception as e:
            logger.error(f'Failed to send internal alert: {str(e)}')

        return Response({
            'detail': 'Client rejected successfully',
            'client': ExistingClientSerializer(client).data
        })

    @action(detail=False, methods=['post'])
    def bulk_approve(self, request):
        """
        Approve multiple client records.

        POST /api/v1/clients/bulk-approve/
        """
        serializer = BulkApprovalSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        client_ids = serializer.validated_data['client_ids']

        # Get all pending clients
        clients = ExistingClient.objects.filter(
            id__in=client_ids,
            approval_status='pending'
        )

        # Update all to approved
        updated_count = clients.update(approval_status='approved')

        # TODO: Create employee accounts for approved clients
        # TODO: Send SMS notifications

        # Send internal alert
        if updated_count > 0:
            try:
                alert_message = f"""
                <p><strong>Bulk Client Approval Completed</strong></p>
                <ul>
                    <li><strong>Clients Approved:</strong> {updated_count}</li>
                    <li><strong>Approved by:</strong> {request.user.get_full_name()}</li>
                </ul>
                """
                send_internal_alert(
                    subject=f'Bulk Approval - {updated_count} clients approved',
                    message=alert_message,
                    alert_type='success'
                )
            except Exception as e:
                logger.error(f'Failed to send internal alert: {str(e)}')

        return Response({
            'detail': f'{updated_count} client(s) approved successfully',
            'approved_count': updated_count
        })


@api_view(['GET'])
@permission_classes([AllowAny])
def download_client_template(request):
    """
    Standalone view to download Excel template for bulk client upload.
    No authentication required.
    
    GET /api/v1/clients/template-download/
    """
    from apps.employers.models import Employer
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Client Data Template
    ws1 = wb.active
    ws1.title = "Client Data"
    
    # Define headers (matching frontend form order exactly)
    headers = [
        'Full Name *',
        'National ID Number *',
        'Mobile Number *',
        'Email Address',
        'Employer Name *',
        'Employee ID',
        'Loan Amount (KES) *',
        'Interest Rate (%) *',
        'Loan Start Date *',
        'Repayment Period (Months) *',
        'Disbursement Date *',
        'Disbursement Method *',
        'Amount Paid to Date (KES)',
        'Loan Status'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write headers with styling
    for col, header in enumerate(headers, start=1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Adjust column widths
    column_widths = [20, 18, 15, 25, 25, 15, 18, 15, 18, 22, 18, 20, 22, 15]
    for col, width in enumerate(column_widths, start=1):
        ws1.column_dimensions[ws1.cell(row=1, column=col).column_letter].width = width
    
    # Add sample data rows
    sample_data_rows = [
        [
            'John Kamau',
            '12345678',
            '254712345678',
            'john@example.com',
            'Safaricom PLC',
            'EMP-1234',
            100000,
            5,
            '2026-01-01',
            6,
            '2026-01-05',
            'mpesa',
            0,
            'Active'
        ],
        [
            'Jane Wanjiku',
            '87654321',
            '254723456789',
            'jane@example.com',
            'Kenya Power',
            'EMP-5678',
            150000,
            5,
            '2026-02-01',
            12,
            '2026-02-05',
            'bank',
            25000,
            'Active'
        ]
    ]
    
    for row_idx, sample_data in enumerate(sample_data_rows, start=2):
        for col, value in enumerate(sample_data, start=1):
            ws1.cell(row=row_idx, column=col, value=value)
    
    # Add instructions row
    ws1.cell(row=4, column=1, value="INSTRUCTIONS:")
    ws1.cell(row=4, column=1).font = Font(bold=True, color="FF0000")
    
    instructions = [
        "1. Fields marked with * are required",
        "2. Use exact employer name from 'Employer Reference' sheet",
        "3. Dates must be in YYYY-MM-DD format (e.g., 2026-01-15)",
        "4. Disbursement Method: mpesa, bank, or cash",
        "5. Loan Status: Active, Fully Paid, Defaulted, or Restructured",
        "6. Delete sample rows and this instruction section before uploading"
    ]
    
    for idx, instruction in enumerate(instructions, start=5):
        ws1.cell(row=idx, column=1, value=instruction)
        ws1.cell(row=idx, column=1).font = Font(italic=True, color="666666")
    
    # Sheet 2: Employer Reference
    ws2 = wb.create_sheet(title="Employer Reference")
    
    # Employer reference headers
    ws2.cell(row=1, column=1, value="Employer Name")
    ws2.cell(row=1, column=2, value="Employer ID (for reference)")
    ws2.cell(row=1, column=1).font = header_font
    ws2.cell(row=1, column=2).font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=1).alignment = header_alignment
    ws2.cell(row=1, column=2).alignment = header_alignment
    
    # Set column widths
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 40
    
    # Add active employers
    employers = Employer.objects.filter(is_active=True).order_by('name')
    for idx, employer in enumerate(employers, start=2):
        ws2.cell(row=idx, column=1, value=employer.name)
        ws2.cell(row=idx, column=2, value=str(employer.id))
    
    # Add note at the bottom
    note_row = len(employers) + 3
    ws2.cell(row=note_row, column=1, value="NOTE: Copy the exact employer name to the 'Client Data' sheet")
    ws2.cell(row=note_row, column=1).font = Font(bold=True, color="0000FF")
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=client_upload_template.xlsx'

    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def validate_bulk_upload(request):
    """
    Standalone view to validate bulk upload file without importing.

    POST /api/v1/clients/validate/
    """
    # Use the viewset's validate_upload method
    viewset = ExistingClientViewSet()
    viewset.request = request
    viewset.format_kwarg = None

    return viewset.validate_upload(request)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_upload_clients(request):
    """
    Standalone view to bulk upload client records.

    POST /api/v1/clients/bulk-upload/
    """
    # Use the viewset's bulk_upload method
    viewset = ExistingClientViewSet()
    viewset.request = request
    viewset.format_kwarg = None

    return viewset.bulk_upload(request)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generate_collection_report(request):
    """
    Generate monthly collection/deduction report in Excel format.

    GET /api/v1/clients/collection-report/

    Query Parameters:
    - employer_id: UUID of employer (required for admin, auto-set for HR)
    - month: Month number (1-12), defaults to current month
    - year: Year (YYYY), defaults to current year

    Response: Excel file download
    """
    from apps.employers.models import Employer
    from datetime import datetime
    from calendar import month_name

    # Get query parameters
    employer_id = request.GET.get('employer_id')
    month = request.GET.get('month', datetime.now().month)
    year = request.GET.get('year', datetime.now().year)

    try:
        month = int(month)
        year = int(year)
    except (ValueError, TypeError):
        return Response(
            {'error': 'Invalid month or year parameter'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Permission check and employer filter
    if request.user.role == 'hr':
        # HR users can only see their own employer
        if not request.user.employer:
            return Response(
                {'error': 'HR user not associated with an employer'},
                status=status.HTTP_400_BAD_REQUEST
            )
        employer = request.user.employer
    elif request.user.role == 'admin':
        # Admin must specify employer
        if not employer_id:
            return Response(
                {'error': 'employer_id parameter is required for admin users'},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            employer = Employer.objects.get(id=employer_id, is_active=True)
        except Employer.DoesNotExist:
            return Response(
                {'error': 'Employer not found or inactive'},
                status=status.HTTP_404_NOT_FOUND
            )
    else:
        return Response(
            {'error': 'Permission denied'},
            status=status.HTTP_403_FORBIDDEN
        )

    # Get all approved clients for this employer with active loans
    clients = ExistingClient.objects.filter(
        employer=employer,
        approval_status='approved',
        loan_status='Active',
        outstanding_balance__gt=0
    ).order_by('full_name')

    if not clients.exists():
        return Response(
            {'error': 'No active clients found for this employer'},
            status=status.HTTP_404_NOT_FOUND
        )

    # Generate Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Deductions"

    # Month name for title
    month_str = month_name[month]
    title = f"{month_str} {year} -Deductions"

    # Row 1: Title (merged across columns)
    ws.merge_cells('B1:D1')
    title_cell = ws['B1']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 2: Headers
    headers = ['Name', 'Amount Borrowed', 'Installment Due']
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, header in enumerate(headers, start=2):  # Start at column B
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_alignment

    # Set column widths
    ws.column_dimensions['A'].width = 5   # Serial number
    ws.column_dimensions['B'].width = 30  # Name
    ws.column_dimensions['C'].width = 18  # Amount Borrowed
    ws.column_dimensions['D'].width = 18  # Installment Due

    # Data rows
    total_borrowed = Decimal('0')
    total_installment = Decimal('0')

    for idx, client in enumerate(clients, start=1):
        row_num = idx + 2  # Start from row 3 (after title and headers)

        # Serial number
        ws.cell(row=row_num, column=1, value=idx)

        # Name
        ws.cell(row=row_num, column=2, value=client.full_name)

        # Amount Borrowed (loan_amount)
        ws.cell(row=row_num, column=3, value=float(client.loan_amount))

        # Installment Due (monthly_deduction)
        ws.cell(row=row_num, column=4, value=float(client.monthly_deduction))

        # Add to totals
        total_borrowed += client.loan_amount
        total_installment += client.monthly_deduction

    # Totals row
    total_row = len(clients) + 3
    ws.cell(row=total_row, column=3, value=float(total_borrowed))
    ws.cell(row=total_row, column=4, value=float(total_installment))

    # Format totals row
    for col in [3, 4]:
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)

    # Format currency columns
    for row in range(3, total_row + 1):
        for col in [3, 4]:  # Amount Borrowed and Installment Due
            cell = ws.cell(row=row, column=col)
            cell.number_format = '#,##0'

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    filename = f"{employer.name}_{month_str}_{year}_Deductions.xlsx"
    filename = filename.replace(' ', '_')

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
